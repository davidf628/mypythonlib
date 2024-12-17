from openpyxl import load_workbook

def load_dictarray(wbfile):

    wb = load_workbook(wbfile)
    ws = wb.active

    dict_array = []
    headers = {}
    
    # Get all the row headers
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value != None:
            headers[value] = col

    # Now get all the student data
    for row in range(2, ws.max_row + 1):
        record = {}
        for header in headers.keys():
            col = headers[header]
            value = ws.cell(row=row, column=col).value
            if value != None:
                record[header] = value
            else:
                record[header] = ''
        dict_array.append(record)

    wb.close()
            
    return dict_array

###############################################################################
# Grabbed from stack overflow to figure out a better estimate on the number of
#  students in the worksheet since ws.max_row often gives an overestimate based
#  on previously deleted data

def find_last_row(ws, column):
    """
    Returns the last row of a given column that has a value in it.
    Empty rows before the last row are allowed.
    ws: a worksheet object
    column: a name, such as 'A'
    """
    selected_column = ws[column]
    # Start at bottom of column
    for row in range(len(selected_column)-1, -1, -1):
        if selected_column[row].value != None:
            return row + 1 # +1 since Excel row numbering starts at 1
    return None # Empty column
